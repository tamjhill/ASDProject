"""takes excel, csv, tsv or txt files retrieved from pubmed, checks if they contain a column re. log fold changes, and saves sheet as csv file"""

import pandas as pd
import os
from openpyxl import load_workbook
import xlrd
import csv
import re

def process_excel_file(file_path):
    """loads excel files into dataframes
    """
    wb = load_workbook(filename=file_path, read_only=True)
    output_dir = os.path.dirname(file_path)
    
    for sheet_name in wb.sheetnames:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        process_dataframe(df, sheet_name, output_dir, file_path)
        

def process_old_file(file_path):
    """loads older-style excel files (.xls) into dataframes
    """
    wb = xlrd.open_workbook(file_path)
    output_dir = os.path.dirname(file_path)
    for sheet in wb.sheets():
        print(f"Processing sheet: {sheet.name}")
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        data = [
            [sheet.cell_value(row, col) for col in range(sheet.ncols)]
            for row in range(1, sheet.nrows)
        ]
        df = pd.DataFrame(data, columns=headers)
        
        process_dataframe(df, sheet.name, output_dir, file_path)


def process_csv_file(file_path):
    """loads csv, tsc or txt friles into dataframes
    """
    output_dir = os.path.dirname(file_path)
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    # Try reading with different settings
    for encoding in ['utf-8', 'iso-8859-1', 'latin1']:
        for delim in ['\t', ',', ';']:  # prioritize tab
            try:
                df = pd.read_csv(file_path, delimiter=delim, encoding=encoding, on_bad_lines='warn')
                if not df.empty:
                    process_dataframe(df, file_name, output_dir, file_path, input_delimiter=delim)
                    return
            except Exception as e:
                print(f"Failed to read {file_path} with delimiter '{delim}' and encoding '{encoding}': {str(e)}")
    
    # If all attempts fail, try reading as plain text
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        print(f"File {file_path} couldn't be processed as CSV. Content preview:")
        print(content[:100])  # Print first 100 characters
    except Exception as e:
        print(f"Failed to read {file_path} as text: {str(e)}")


def process_dataframe(df, sheet_name, output_dir, file_path, input_delimiter='\t'):
    """processes dataframes to assess if the data relates to gene expression - looks for "log fold change" or similar
    in column titles
    """
    df.columns = df.columns.astype(str)
    log_fold_col = None
    for col in df.columns:
        if any(phrase in re.sub(r'[_\s-]', '', col.lower()) for phrase in ['logfoldchange', 'logfold', 'logfold2', 'lf', 
                                                                           'expression', 'enrichment', 'logfc', 'foldchange', 'fc', 
                                                                           'log2', 'lf2', 'lfc', 'log2fc', 'log', 'fold']):
            log_fold_col = col
            break
    
    # If matching column is found, save sheet as CSV
    if log_fold_col:
        # remove characters not appropriate for filenames
        replacement_chars  = {" " : "",
                              "<" : "lessthan",
                              ">" : "morethan",
                              "." : "",
                              ":" : "",
                              "/" : "",
                              "?" : "",
                              "*" : "",
                              "&" : "" }
        for old, new in replacement_chars.items():
            sheet_name = sheet_name.replace(old, new)
        newfile = 'expdata_' + sheet_name
        output_file = os.path.join(output_dir, f"{newfile}.csv")

        if os.path.exists(output_file):
            original_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_file = os.path.join(output_dir, f"{newfile}_{original_filename}.csv")

        if input_delimiter == '\t':
            df.to_csv(output_file, index=False, sep=',')
        else:
            df.to_csv(output_file, index=False)
        print(f"Saved {sheet_name} as CSV: {output_file}")
    else:
        print(f"Skipped {sheet_name} in {file_path}: No 'log fold change' column found")


def process_data_folder(data_folder):
    """ Walks through all subdirectories, processes files found according to their extension
    """
    for root, dirs, files in os.walk(data_folder):
        for file in files:
            file_path = os.path.join(root, file)
            print(f"Processing file: {file_path}")
            if file.lower().endswith('.xlsx'):
                process_excel_file(file_path)
            elif file.lower().endswith('.xls'):
                process_old_file(file_path)
            elif file.lower().endswith(('.csv', '.tsv', '.txt')):
                process_csv_file(file_path)


data_folder = "data\\supp_data\\"
process_data_folder(data_folder)